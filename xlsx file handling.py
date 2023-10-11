import openpyxl
'''Accessing and reading an existing Excel workbook'''
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']  # Replace 'Sheet1' with the name of your sheet
#
# # Read data from a specific cell:
# cell_value = sheet['A1'].value
#
# # You can iterate through rows and columns to read multiple values:
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     name, age = row
#     print(f"Name: {name}, Age: {age}")

'''Writing to a workbook'''
# Open an existing workbook or create a new one:
wb = openpyxl.Workbook()

# Access a specific sheet in the workbook:
sheet = wb.active  # This selects the active sheet, usually the first one

# Write data to specific cells
sheet['A1'] = "Name"
sheet['B1'] = "Age"
sheet['A2'] = "John"
sheet['B2'] = 25

# Saving and closing the workbook
wb.save('output.xlsx')
wb.close()